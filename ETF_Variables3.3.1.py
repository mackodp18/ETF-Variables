#---------------------------------------------------------------------------------------
#Import Libraries
import os
import openpyxl
import csv
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl import workbook
import xlsxwriter
import zipfile
from zipfile import ZipFile, ZIP_DEFLATED
#-------------------------------------------------------------------------------------------
#PROCESS FUNCTIONS TAKING USER INPUT AND ADDING TO EXCEL SHEET
def CeramicMaterial_decleration():
    i = 1
    while i <= int(SN):
        print('SN:' + str(i))
        sheet_Materials['A' + str(i + 2)] = (str('SN') + str(i))
        sheet_Materials['B' + str(2)].value = str('Ceramic Material')
        print('What ceramic Material are you using?')
        Ceramic = input()
        sheet_Materials['B' + str(i + 2)].value = str(Ceramic)
        sheet_Materials['B' + str(2)].value = str('Ceramic Material')
        print('What is the surface roughness?')
        Ceramic_SR = input()
        sheet_Materials['C' + str(i + 2)].value = str(Ceramic_SR)
        sheet_Materials['C' + str(2)].value = str('CSurface Rouhgness (Ra)')
        print('What is the size of the Ceramic?')
        Size_Ceramic = input()
        sheet_Materials['D' + str(i + 2)].value = str(Size_Ceramic)
        sheet_Materials['D' + str(2)].value = str('Ceramic size (inches)')
        print('What is the thickness of the Ceramic?')
        th_Ceramic = input()
        sheet_Materials['E' + str(i + 2)].value = str(th_Ceramic)
        sheet_Materials['E' + str(2)].value = str('Ceramic Thickness (inches)')
        i = i + 1      

def Additional_Ceramic_Material():
    i = int(SN) + 1
    while i <= int(int(SN) + int(SN_added)):
        print('SN:' + str(i))
        ws['A' + str(i + int(SN))].value = (str('SN') + str(i))
        print('What ceramic Material are you using?')
        Ceramic = input()
        ws['B' + str(i + int(SN))].value = str(Ceramic)
        print('What is the surface roughness?')
        Ceramic_SR = input()
        ws['C' + str(i + int(SN))].value = str(Ceramic_SR)
        print('What is the size of the Ceramic?')
        Size_Ceramic = input()
        ws['D' + str(i + int(SN))].value = str(Size_Ceramic)
        print('What is the thickness of the Ceramic?')
        th_Ceramic = input()
        ws['E' + str(i + int(SN))].value = str(th_Ceramic)
        i = i + 1     

def Print_process():
    print('How many Layers will you be printing?')
    Layer_numbers = input()
    i = 1
    SNrow = 3
    while i <= int(SN):
        x = 1
        row = SNrow
        while x <= int(Layer_numbers):
            sheet_Print['A' + str(row)].value = (str('SN') + str(i) + str(' Layer# ') + str(x))
            sheet_Bake['A' + str(row)].value = (str('SN') + str(i) + str(' Layer# ') + str(x))
            sheet_Fire['A' + str(row)].value = (str('SN') + str(i) + str(' Layer# ') + str(x))
            print('SN:' + str(i) + str(' Layer#') + str(x))
            print('What Layer are you printing?')
            Layer = input()
            sheet_Print['B' + str(2)].value = str('Layer')
            sheet_Print['B' + str(row)].value = str(Layer)
            print('What material are you using (ex: Dupont 5771, Gold)')
            Paste = input()
            sheet_Print['C' + str(2)].value = str('Material')
            sheet_Print['C' + str(row)].value = str(Paste)
            print('What is your screen size?')
            Screen_size = input()
            sheet_Print['D' + str(2)].value = str('Screen Size')
            sheet_Print['D' + str(row)].value = str(Screen_size)
            print('What is the mesh size?')
            Mesh_size = input()
            sheet_Print['E' + str(2)].value = str('Mesh Size')
            sheet_Print['E' + str(row)].value = str(Mesh_size)
            print('What is the paste viscosity?')
            Viscosity = input()
            sheet_Print['F' + str(2)].value = str('Viscosity')
            sheet_Print['F' + str(row)].value = str(Viscosity)
            print('What is the downstop?')
            downstop = input()
            sheet_Print['G' + str(2)].value = str('Downstop')
            sheet_Print['G' + str(row)].value = str(downstop)
            print('What is the pressure?')
            Pressure = input()
            sheet_Print['H' + str(2)].value = str('Pressure')
            sheet_Print['H' + str(row)].value = str(Pressure)
            print('What the print speed?')
            Print_Speed = input()
            sheet_Print['I' + str(2)].value = str('Print Speed')
            sheet_Print['I' + str(row)].value = str(Print_Speed)
            print('What is the snapoff?')
            Snapoff = input()
            sheet_Print['J' + str(2)].value = str('Snapoff')
            sheet_Print['J' + str(row)].value = str(Snapoff)
            print('What Durometer are you using?')
            Durometer = input()
            sheet_Print['K' + str(2)].value = str('Durometer')
            sheet_Print['K' + str(row)].value = str(Durometer)
            print('What is the wet thickness?')
            Wet_Thickness = input()
            sheet_Print['L' + str(2)].value = str('Wet Thickness')
            sheet_Print['L' + str(row)].value = str(Wet_Thickness)
            print('What is the temperature of the baking oven (Celsius)?')
            Temp_Oven = input()
            sheet_Bake['B' + str(2)].value = str('Temperature (C)')
            sheet_Bake['B' + str(row)].value = str(Temp_Oven)
            print('How long were the parts in the oven (minutes)?')
            Time_Oven = input()
            sheet_Bake['C' + str(2)].value = str('Time (Minutes)')
            sheet_Bake['C' + str((row))].value = str(Time_Oven)
            print('What is the Dried thickness?')
            Dried_th = input()
            sheet_Bake['D' + str(2)].value = str('Dried Thickness')
            sheet_Bake['D' + str(row)].value = str(Dried_th)
            print('How many zones is the furnace?')
            Number_Zones = input()
            print('What is the input temperature of the Furnace (Celsius)?')
            Input_Temp_Furnace = input()
            sheet_Fire['B' + str(2)].value = str('Input Temperature (C)')
            sheet_Fire['B' + str(row)].value = str(Input_Temp_Furnace)
            print('What is the Dwell time of the firing profile (minutes)?')
            Dwell_Time = input()
            sheet_Fire['C' + str(2)].value = str('Dwell Time (Minutes)')
            sheet_Fire['C' + str(row)].value = str(Dwell_Time)
            print('What is the Peak temperature of the Furnace (Celsius)?')
            Peak_Temp_Furnace = input()
            sheet_Fire['D' + str(2)].value = str('Peak Temperature (C)')
            sheet_Fire['D' + str(row)].value = str(Peak_Temp_Furnace)
            print('What is the ramp time (minutes)?')
            Ramp_Time = input()
            sheet_Fire['E' + str(2)].value = str('Ramp Time (minutes)')
            sheet_Fire['E' + str(row)].value = str(Ramp_Time)
            print('What is the cool down time (minutes)?')
            CoolDown_Time = input()
            sheet_Fire['F' + str(2)].value = str('Cool Down Time (minutes)')
            sheet_Fire['F' + str(row)].value = str(CoolDown_Time)
            print('What is the fired thickness?')
            Fired_Thickness = input()
            sheet_Fire['G' + str(2)].value = str('Fired Thickness')
            sheet_Fire['G' + str(row)].value = str(Fired_Thickness) 
            x = x + 1
            row = row + 1 
        i = i + 1
        SNrow = SNrow + int(Layer_numbers)

def Additional_Print_Parameters():
    print('How many Layers will you be printing?')
    Layer_numbers = input()
    StartRow = int(SN) * int(Layer_numbers)
    SNrow = StartRow + 3
    z = int(SN) + 1
    while int(z) <= int(int(SN) + int(SN_added)):
        x = 1
        row = SNrow
        while x <= int(Layer_numbers):
            ws['A' + str(row)].value = (str('SN') + str(z) + str(' Layer# ') + str(x))
            ws2['A' + str(row)].value = (str('SN') + str(z) + str(' Layer# ') + str(x))
            ws3['A' + str(row)].value = (str('SN') + str(z) + str(' Layer# ') + str(x))
            print('SN:' + str(z) + str(' Layer#') + str(x))
            print('What Layer are you printing?')
            Layer = input()
            ws['B' + str(2)].value = str('Layer')
            ws['B' + str(row)].value = str(Layer)
            print('What material are you using (ex: Dupont 5771, Gold)')
            Paste = input()
            ws['C' + str(2)].value = str('Material')
            ws['C' + str(row)].value = str(Paste)
            print('What is your screen size?')
            Screen_size = input()
            ws['D' + str(2)].value = str('Screen Size')
            ws['D' + str(row)].value = str(Screen_size)
            print('What is the mesh size?')
            Mesh_size = input()
            ws['E' + str(2)].value = str('Mesh Size')
            ws['E' + str(row)].value = str(Mesh_size)
            print('What is the paste viscosity?')
            Viscosity = input()
            ws['F' + str(2)].value = str('Viscosity')
            ws['F' + str(row)].value = str(Viscosity)
            print('What is the downstop?')
            downstop = input()
            ws['G' + str(2)].value = str('Downstop')
            ws['G' + str(row)].value = str(downstop)
            print('What is the pressure?')
            Pressure = input()
            ws['H' + str(2)].value = str('Pressure')
            ws['H' + str(row)].value = str(Pressure)
            print('What the print speed?')
            Print_Speed = input()
            ws['I' + str(2)].value = str('Print Speed')
            ws['I' + str(row)].value = str(Print_Speed)
            print('What is the snapoff?')
            Snapoff = input()
            ws['J' + str(2)].value = str('Snapoff')
            ws['J' + str(row)].value = str(Snapoff)
            print('What Durometer are you using?')
            Durometer = input()
            ws['K' + str(2)].value = str('Durometer')
            ws['K' + str(row)].value = str(Durometer)
            print('What is the wet thickness?')
            Wet_Thickness = input()
            ws['L' + str(2)].value = str('Wet Thickness')
            ws['L' + str(row)].value = str(Wet_Thickness)
            print('What is the temperature of the baking oven (Celsius)?')
            Temp_Oven = input()
            ws2['B' + str(2)].value = str('Temperature (C)')
            ws2['B' + str(row)].value = str(Temp_Oven)
            print('How long were the parts in the oven (minutes)?')
            Time_Oven = input()
            ws2['C' + str(2)].value = str('Time (Minutes)')
            ws2['C' + str((row))].value = str(Time_Oven)
            print('What is the Dried thickness?')
            Dried_th = input()
            ws2['D' + str(2)].value = str('Dried Thickness')
            ws2['D' + str(row)].value = str(Dried_th)
            print('How many zones is the furnace?')
            Number_Zones = input()
            print('What is the input temperature of the Furnace (Celsius)?')
            Input_Temp_Furnace = input()
            ws3['B' + str(2)].value = str('Input Temperature (C)')
            ws3['B' + str(row)].value = str(Input_Temp_Furnace)
            print('What is the Dwell time of the firing profile (minutes)?')
            Dwell_Time = input()
            ws3['C' + str(2)].value = str('Dwell Time (Minutes)')
            ws3['C' + str(row)].value = str(Dwell_Time)
            print('What is the Peak temperature of the Furnace (Celsius)?')
            Peak_Temp_Furnace = input()
            ws3['D' + str(2)].value = str('Peak Temperature (C)')
            ws3['D' + str(row)].value = str(Peak_Temp_Furnace)
            print('What is the ramp time (minutes)?')
            Ramp_Time = input()
            ws3['E' + str(2)].value = str('Ramp Time (minutes)')
            ws3['E' + str(row)].value = str(Ramp_Time)
            print('What is the cool down time (minutes)?')
            CoolDown_Time = input()
            ws3['F' + str(2)].value = str('Cool Down Time (minutes)')
            ws3['F' + str(row)].value = str(CoolDown_Time)
            print('What is the fired thickness?')
            Fired_Thickness = input()
            ws3['G' + str(2)].value = str('Fired Thickness')
            ws3['G' + str(row)].value = str(Fired_Thickness)   
            x = x + 1
            row = row + 1
        z = z + 1
        SNrow = SNrow + int(Layer_numbers)



def Bake_process():
    i = 1
    while i <= int(SN):
        print('SN:' + str(i))
        sheet_Bake['A' + str(i + 2)].value = (str('SN') + str(i))
        print('What is the temperature of the baking oven (Celsius)?')
        Temp_Oven = input()
        sheet_Bake['B' + str(2)].value = str('Temperature (C)')
        sheet_Bake['B' + str(i + 2)].value = str(Temp_Oven)
        print('How long were the parts in the oven (minutes)?')
        Time_Oven = input()
        sheet_Bake['C' + str(2)].value = str('Time (Minutes)')
        sheet_Bake['C' + str(i + 2)].value = str(Time_Oven)
        print('What is the Dried thickness?')
        Dried_th = input()
        sheet_Bake['D' + str(2)].value = str('Dried Thickness')
        sheet_Bake['D' + str(i + 2)].value = str(Dried_th)
        i = i + 1

def Fire_Process():
    print('How many zones is the furnace?')
    Number_Zones = input()
    i = 1
    while i <= int(SN):
        print('SN:' + str(i))
        sheet_Fire['A' + str(i + 2)].value = (str('SN') + str(i))
        print('What is the input temperature of the Furnace (Celsius)?')
        Input_Temp_Furnace = input()
        sheet_Fire['B' + str(2)].value = str('Input Temperature (C)')
        sheet_Fire['B' + str(i + 2)].value = str(Input_Temp_Furnace)
        print('What is the Dwell time of the firing profile (minutes)?')
        Dwell_Time = input()
        sheet_Fire['C' + str(2)].value = str('Dwell Time (Minutes)')
        sheet_Fire['C' + str(i + 2)].value = str(Dwell_Time)
        print('What is the Peak temperature of the Furnace (Celsius)?')
        Peak_Temp_Furnace = input()
        sheet_Fire['D' + str(2)].value = str('Peak Temperature (C)')
        sheet_Fire['D' + str(i + 2)].value = str(Peak_Temp_Furnace)
        print('What is the ramp time (minutes)?')
        Ramp_Time = input()
        sheet_Fire['E' + str(2)].value = str('Ramp Time (minutes)')
        sheet_Fire['E' + str(i + 2)].value = str(Ramp_Time)
        print('What is the cool down time (minutes)?')
        CoolDown_Time = input()
        sheet_Fire['F' + str(2)].value = str('Cool Down Time (minutes)')
        sheet_Fire['F' + str(i + 2)].value = str(CoolDown_Time)   
        n = 1
        while n <= int(Number_Zones):
            Length_iter = int(SN) + int(i) + int(n * int(Number_Zones))
            sheet_Fire['A' + str(Length_iter)].value = (str('SN') + str(i) + 'Zone#' + str(n))
            print('What is the temperature in zone ' + str(n) + '?')
            Zone_Temp = input()
            sheet_Fire['B' + str(Length_iter)].value = str(Zone_Temp)
            n = n + 1
        i = i + 1

def Etch_Process():
    i = 1
    while i <= int(SN):
        print('SN:' + str(i))
        sheet_Etch['A' + str(i + 2)].value = (str('SN') + str(i))
        print('What is the etch Factor?')
        EF = input()
        sheet_Etch['B' + str(2)].value = str('pEtch Factor')
        sheet_Etch['B' + str(i + 2)].value = str(EF)
        print('How many passes of photoresist?')
        Photoresist_passes = input()
        sheet_Etch['C' + str(2)].value = str('passes of photoresist')
        sheet_Etch['C' + str(i + 2)].value = str(Photoresist_passes)
        print('What is the panel setup in the photoresist Chamber? (Left -> right x Top -> Down)')
        Photoresist_setup = input()
        sheet_Etch['D' + str(2)].value = str('Setup')
        sheet_Etch['D' + str(i + 2)].value = str(Photoresist_setup)
        print('What is the thickness of the photoresist?')
        Photoresist_thickness = input()
        sheet_Etch['E' + str(2)].value = str('Photoresist thickness')
        sheet_Etch['E' + str(i + 2)].value = str(Photoresist_thickness)
        print('What is the temperature of the baking oven (Celsius)?')
        Temp_Oven_etch = input()
        sheet_Etch['F' + str(2)].value = str('Temperature (C)')
        sheet_Etch['F' + str(i + 2)].value = str(Temp_Oven_etch)
        print('How long were the parts in the oven (minutes)?')
        Time_Oven_etch = input()
        sheet_Etch['G' + str(2)].value = str('Time (Minutes)')
        sheet_Etch['G' + str(i + 2)].value = str(Time_Oven_etch)
        print('What is the Dried thickness?')
        Dried_th_etch = input()
        sheet_Etch['H' + str(2)].value = str('Dried Thickness')
        sheet_Etch['H' + str(i + 2)].value = str(Dried_th_etch)
        print('What is the exposure artwork? (Glass or Mylar)')
        Artwork = input()
        sheet_Etch['I' + str(2)].value = str('Artwork')
        sheet_Etch['I' + str(i + 2)].value = str(Artwork)
        print('What is the bulb power? (mW)')
        bulb_power = input()
        sheet_Etch['J' + str(2)].value = str('bulb power')
        sheet_Etch['J' + str(i + 2)].value = str(bulb_power)
        print('What is the exposure time? (s)')
        exposure_time = input()
        sheet_Etch['K' + str(2)].value = str('exposure time')
        sheet_Etch['K' + str(i + 2)].value = str(exposure_time)
        print('What is the develop temperature? (Celsius)')
        develop_temp = input()
        sheet_Etch['L' + str(2)].value = str('develop temp (C)')
        sheet_Etch['L' + str(i + 2)].value = str(develop_temp)
        print('What is the develop time? (minutes:s)')
        develop_time = input()
        sheet_Etch['M' + str(2)].value = str('develop time')
        sheet_Etch['M' + str(i + 2)].value = str(develop_time)
        print('What is the etcher temperature? (Celsius)')
        etcher_temp = input()
        sheet_Etch['N' + str(2)].value = str('etcher temp (C)')
        sheet_Etch['N' + str(i + 2)].value = str(etcher_temp)
        print('What is the etch time? (minutes:s)')
        etcher_time = input()
        sheet_Etch['O' + str(2)].value = str('etch time')
        sheet_Etch['O' + str(i + 2)].value = str(etcher_time)
        i = i + 1
#Main
while True:
    print('Will you be adding to a current excel workbook? (yes or no)')
    command = input()
    if command in ('yes', 'no'):
        break
    else:
        print('Please enter yes or no.')
print('What is the part number?')
PN = input()
print('What is the name of your experiment?')
FileName = input()
#Create Worksheets
wb = openpyxl.Workbook()
sheet_Materials = wb.create_sheet(str(FileName) + str(' Material')) 
sheet_Print = wb.create_sheet(str(FileName) + str(' Print Parameters'))
sheet_Bake = wb.create_sheet(str(FileName) + str(' Bake Parameters'))
sheet_Fire = wb.create_sheet(str(FileName) + str(' Fire Parameters'))
sheet_Etch = wb.create_sheet(str(FileName) + str(' Etch Parameters'))
#Main Code
if command == str('yes'):
    print('please add the filepth where the excel file can be found')
    filepath = input()
    os.chdir(filepath)
    print('What will you be adding to the worksheet?' '\n' '1 = Additional Ceramic' '\n' '2 = Additional Ceramic and Print Parameters' '\n' '3 = Additional Print Parameters to current Ceramics' '\n' '4 = Add Etch Process' )
    process = input()
    print('How many Ceramics are already in this study?')
    SN = input()
    print('Please enter a number.')
    print('How many Ceramics will you be adding?')
    SN_added = input()
    n = 0
    if process == str(1):
        n = 1
        cleanpath = os.path.abspath(filepath)
        wb = load_workbook((FileName + str('.xlsx')))
        sheets = wb.sheetnames
        ws = wb[sheets[n]]
        Additional_Ceramic_Material()
        wb.save(FileName + str('.xlsx'))
        wb.save(FileName + str('.xcsv'))
    elif process == str(2):
        n = 2
        cleanpath = os.path.abspath(filepath)
        wb = load_workbook((FileName + str('.xlsx')))
        sheets = wb.sheetnames
        ws = wb[sheets[n]]
        ws2 = wb[sheets[n + 1]]
        ws3 = wb[sheets[n + 2]]
        Additional_Print_Parameters()
        wb.save(FileName + str('.xlsx'))
        wb.save(FileName + str('.csv'))
    elif process == str(3):
        n = 2
        cleanpath = os.path.abspath(filepath)
        wb = load_workbook((FileName + str('.xlsx')))
        sheets = wb.sheetnames
        sheet_Print = wb[sheets[n]]
        sheet_Bake = wb[sheets[n + 1]]
        sheet_Fire = wb[sheets[n + 2]]
        Print_process()
        wb.save(FileName + str('.xlsx'))
        wb.save(FileName + str('.csv')) 
    elif process == str(4):
        n = 2
        cleanpath = os.path.abspath(filepath)
        wb = load_workbook((FileName + str('.xlsx')))
        sheets = wb.sheetnames
        sheet_Etch = wb[sheets[n + 3]]
        Etch_Process()
        wb.save(FileName + str('.xlsx'))
        wb.save(FileName + str('.csv')) 
if command == str('no'):
    print('Please add the filepath where you would like this document placed.')
    Directory = input()
    os.chdir(Directory)   
    os.mkdir(FileName)
    os.chdir(Directory + '/' + FileName)
    os.mkdir('Results')
    print('What process is being performed?' '\n' '1 = Print' '\n' '2 = Etch' '\n' '3 = Bake' '\n' '4 = Fire' )
    process = input()
    print('How many Ceramics are you using?')
    SN = input()
    if process == str(1):
        CeramicMaterial_decleration()
        Print_process()
    elif process == str(2):
        Etch_Process()  
    elif process == str(3):
        Bake_process()
    elif process == str(4):
        Fire_Process()
    else:
        print('Please enter a number')   
    wb.save(FileName + '.xlsx')
    wb.save(FileName + '.csv')




